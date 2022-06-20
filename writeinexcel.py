from openpyxl import Workbook
import scrapping

from datetime import date

book = Workbook()
sheet = book.active
json_data=(scrapping.get_jason_data())
for i in range(len(json_data)) :
    sub_obj = json_data[i]
    print("sub",sub_obj)
    if i == 0 :
        keys = list(sub_obj.keys())
        print("keys",keys)
        for k in range(len(keys)) :
			# row or column index start from 1
            sheet.cell(row = (i + 1), column = (k + 1), value = keys[k]);
    for j in range(len(keys)) :
        sheet.cell(row = (i + 2), column = (j + 1), value = sub_obj[keys[j]]);

# for list_row_value in scrapping.get_jason_data():
#     sheet.append(list_row_value)




book.save(f'{date.today()}Share_Price.xlsx')
